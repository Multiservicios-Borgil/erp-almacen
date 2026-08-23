import re
import os

# 1. Leer el archivo original completo que pusiste en main.txt
with open('main.txt', 'r', encoding='utf-8') as f:
    content = f.read()

# 2. Corregir el error de TemplateResponse de forma masiva
# Este patron busca el formato antiguo y lo pasa al nuevo formato de FastAPI
def fix_templates(text):
    # Primero el caso multilineal
    text = re.sub(r'templates\.TemplateResponse\(\s*([\"\'].*?[\"\']),\s*(\{.*?request.*?\})\s*\)', 
                  r'templates.TemplateResponse(request=request, name=\1, context=\2)', text, flags=re.DOTALL)
    # Luego el caso lineal simple
    text = re.sub(r'templates\.TemplateResponse\(([\"\'].*?[\"\']),\s*(\{.*?request.*?\})\)', 
                  r'templates.TemplateResponse(request=request, name=\1, context=\2)', text)
    return text

content = fix_templates(content)

# 3. Limpiar duplicados de importaciones si los hay
content = content.replace('from .models import Evento, Venta', '')

# 4. Añadir la lógica de Amazon Truck (Camión 2) al final
if 'procesar_camion_2' not in content:
    logic = """
# --- AMAZON TRUCK LOGIC ---
import openpyxl, io, uuid, qrcode

@app.get("/importar_amazon", response_class=HTMLResponse)
def importar_amazon_form(request: Request):
    return templates.TemplateResponse(request=request, name="importar_amazon.html")

@app.post("/importar_amazon")
async def procesar_amazon(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    sheet = wb.active
    mapa = {"lavadora secadora": 12, "lavasecadora": 12, "lavadora": 1, "frigorifico": 2, "arcon": 11}
    prefijos = {1:"LAV", 2:"FRI", 11:"ARC", 12:"LSEC"}
    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[2]: continue
        desc = str(row[2]).lower()
        fid = next((v for k,v in mapa.items() if k in desc), None)
        id_gen = f"{prefijos.get(fid, 'AMZ')}-{str(uuid.uuid4())[:4].upper()}"
        item = Item(id=id_gen, familia_id=fid, nombre_pieza=str(row[2])[:100], numero_serie=str(row[3]), 
                    estado_actual='PENDIENTE_CLASIFICAR', origen='AMAZON', camion=2)
        db.add(item)
        items.append(item)
        qr = qrcode.make(f"{request.base_url}item/{id_gen}")
        qr.save(f"app/static/{id_gen}.png")
    db.commit()
    return templates.TemplateResponse(request=request, name="etiquetas_lote.html", context={"items": items})

@app.get("/procesar_camion_2")
async def procesar_camion_2(request: Request, db: Session = Depends(get_db)):
    path = "Copia de ELECTRO ILLUECA 2.xlsx"
    if not os.path.exists(path): return HTMLResponse("Error: Excel no encontrado")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    mapa = {"lavadora secadora": 12, "lavadora": 1, "frigorifico": 2, "arcon": 11}
    prefijos = {1:"LAV", 2:"FRI", 11:"ARC", 12:"LSEC"}
    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[2]: continue
        desc = str(row[2]).lower()
        fid = next((v for k,v in mapa.items() if k in desc), None)
        id_gen = f"{prefijos.get(fid, 'AMZ')}-{str(uuid.uuid4())[:4].upper()}"
        item = Item(id=id_gen, familia_id=fid, nombre_pieza=str(row[2])[:100], numero_serie=str(row[3]), 
                    estado_actual='PENDIENTE_CLASIFICAR', origen='AMAZON', camion=2)
        db.add(item)
        items.append(item)
        qr = qrcode.make(f"{request.base_url}item/{id_gen}")
        qr.save(f"app/static/{id_gen}.png")
    db.commit()
    return templates.TemplateResponse(request=request, name="etiquetas_lote.html", context={"items": items})
"""
    content += logic

# 5. Escribir el archivo final en app/main.py
with open('app/main.py', 'w', encoding='utf-8') as f:
    f.write(content)

print("CORRECCION EXITOSA: app/main.py ha sido restaurado y reparado.")
