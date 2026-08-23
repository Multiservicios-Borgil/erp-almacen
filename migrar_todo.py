import re
import os

# Leemos el archivo original de 1285 lineas que el usuario recuperó
path_original = 'main.py'
with open(path_original, 'r', encoding='utf-8') as f:
    content = f.read()

# 1. Corregir las importaciones (quitar Evento y Venta que no existen en el nuevo models.py)
content = content.replace('from .models import Evento, Venta', '')
content = content.replace('from .models import Base, Item, Familia, HistorialDiagnostico', 'from .models import Base, Item, Familia, Imagen, HistorialDiagnostico')

# 2. Corregir el formato de TemplateResponse (el error unhashable type: dict)
# Buscamos el formato viejo: templates.TemplateResponse("archivo.html", {"request": request, ...})
# Y lo pasamos al nuevo: templates.TemplateResponse(request=request, name="archivo.html", context={...})

def fix_templates(text):
    # Caso multilineal
    text = re.sub(r'templates\.TemplateResponse\(\s*([\"\'].*?[\"\']),\s*(\{.*?request.*?\})\s*\)', 
                  r'templates.TemplateResponse(request=request, name=\1, context=\2)', text, flags=re.DOTALL)
    # Caso lineal simple
    text = re.sub(r'templates\.TemplateResponse\(([\"\'].*?[\"\']),\s*(\{.*?request.*?\})\)', 
                  r'templates.TemplateResponse(request=request, name=\1, context=\2)', text)
    return text

content = fix_templates(content)

# 3. Añadir la lógica de Amazon al final (si no está ya)
if 'procesar_camion_2' not in content:
    amazon_logic = """
# ---------------- IMPORTAR AMAZON (NUEVO) ----------------
import openpyxl

@app.get("/importar_amazon", response_class=HTMLResponse)
def importar_amazon_form(request: Request):
    return templates.TemplateResponse(request=request, name="importar_amazon.html")

@app.post("/importar_amazon")
async def procesar_amazon(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    sheet = wb.active
    mapa_familias = {
        "lavadora secadora": 12, "lavasecadora": 12, "lavadora": 1, "secadora": 3,
        "frigorifico": 2, "frigo": 2, "vinoteca": 2, "vino": 2,
        "lavavajillas": 4, "horno": 5, "microondas": 6, "aire": 7,
        "termo": 8, "vitro": 9, "campana": 10, "arcon": 11, "congelador": 11
    }
    prefijos = {1: "LAV", 2: "FRI", 3: "SEC", 4: "LAVV", 5: "HOR", 6: "MIC", 7: "AIRE", 8: "TER", 9: "VIT", 10: "CAM", 11: "ARC", 12: "LSEC"}
    items_creados = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[2]: continue
        desc = str(row[2]).lower()
        familia_id = None
        for clave, fid in mapa_familias.items():
            if clave in desc:
                familia_id = fid
                break
        prefijo = prefijos.get(familia_id, "AMZ")
        nuevo_id = f"{prefijo}-{str(uuid.uuid4())[:4].upper()}"
        item = Item(id=nuevo_id, familia_id=familia_id, nombre_pieza=str(row[2])[:100], numero_serie=str(row[3]),
                    estado_actual="PENDIENTE_CLASIFICAR", origen="AMAZON_TRUCK", en_stock=True, camion=2)
        db.add(item)
        items_creados.append(item)
        url = f"{request.base_url}item/{nuevo_id}"
        qr = qrcode.make(url)
        os.makedirs("app/static", exist_ok=True)
        qr.save(f"app/static/{nuevo_id}.png")
    db.commit()
    return templates.TemplateResponse(request=request, name="etiquetas_lote.html", context={"items": items_creados})

@app.get("/procesar_camion_2")
async def procesar_camion_2(request: Request, db: Session = Depends(get_db)):
    file_path = "Copia de ELECTRO ILLUECA 2.xlsx"
    if not os.path.exists(file_path):
        return HTMLResponse("<h2>Error: No se encuentra el archivo excel en la carpeta raíz.</h2>")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    mapa_familias = {
        "lavadora secadora": 12, "lavasecadora": 12, "lavadora": 1, "secadora": 3,
        "frigorifico": 2, "frigo": 2, "vinoteca": 2, "vino": 2,
        "lavavajillas": 4, "horno": 5, "microondas": 6, "aire": 7,
        "termo": 8, "vitro": 9, "campana": 10, "arcon": 11, "congelador": 11
    }
    prefijos = {1: "LAV", 2: "FRI", 3: "SEC", 4: "LAVV", 5: "HOR", 6: "MIC", 7: "AIRE", 8: "TER", 9: "VIT", 10: "CAM", 11: "ARC", 12: "LSEC"}
    items_creados = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[2]: continue
        desc = str(row[2]).lower()
        familia_id = None
        for clave, fid in mapa_familias.items():
            if clave in desc:
                familia_id = fid
                break
        prefijo = prefijos.get(familia_id, "AMZ")
        nuevo_id = f"{prefijo}-{str(uuid.uuid4())[:4].upper()}"
        item = Item(id=nuevo_id, familia_id=familia_id, nombre_pieza=str(row[2])[:100], numero_serie=str(row[3]),
                    estado_actual="PENDIENTE_CLASIFICAR", origen="AMAZON_TRUCK", en_stock=True, camion=2)
        db.add(item)
        items_creados.append(item)
        url = f"{request.base_url}item/{nuevo_id}"
        qr = qrcode.make(url)
        qr.save(f"app/static/{nuevo_id}.png")
    db.commit()
    return templates.TemplateResponse(request=request, name="etiquetas_lote.html", context={"items": items_creados})
"""
    content += amazon_logic

# 4. Asegurar familias nuevas en la lista inicial
if 'Arcón frigorífico' not in content:
    content = content.replace('"Campana extractora",', '"Campana extractora",\n    "Arcón frigorífico",\n    "Lavadora-Secadora",')

with open('app/main.py', 'w', encoding='utf-8') as f:
    f.write(content)
print("Migración completa finalizada con éxito.")
